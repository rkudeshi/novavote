#!/usr/bin/env python3
"""Parse Fairfax's Final AB Daily Report workbooks into the NovaVote CSVs.

The county publishes end-of-cycle reports as .xlsx (the daily in-cycle
reports are PDFs — see parse_report.py for those). Two layouts appear:

  A. 2021-2023 — one column per early-voting site, a `Total` row carrying
     the published grand total, and one row per date.
  B. 2020      — every site is split into congressional-district
     sub-columns (CD 8 / CD 10 / CD 11 / Herndon) with a `Total` column
     closing each block. The site name sits in the row *above* the
     sub-headers, so a block runs from one site name to the next.

Self-checking by construction: each sheet prints its own grand total, and
every daily column is summed and asserted against it. A misparse fails
the run instead of writing plausible-looking numbers — which matters more
here than usual, because nobody is going to re-audit a 2020 spreadsheet
by hand.
"""

import argparse
import csv
import datetime as dt
import re
import sys
from pathlib import Path

import openpyxl

OUT = Path("data/parsed")

# Every spelling the county has used for a site, mapped to our key.
SITE_KEYS = {
    "government center": "government_center",
    "gov't center": "government_center",
    "govt center": "government_center",
    "mt. vernon": "mt_vernon",
    "mt vernon": "mt_vernon",
    "mount vernon": "mt_vernon",
    "north county": "north_county",
    "burke": "burke",
    "centreville": "centreville",
    "franconia": "franconia",
    "great falls": "great_falls",
    "herndon": "herndon_fortnightly",
    "herndon fortnightly": "herndon_fortnightly",
    "lorton": "lorton",
    "laurel hill": "laurel_hill",
    "gerry hyland": "gerry_hyland",
    "mason": "mason",
    "mclean": "mclean",
    "providence": "providence",
    "jim scott": "jim_scott",
    "sully": "sully",
    "thomas jefferson": "thomas_jefferson",
    "tysons pimmit": "tysons_pimmit",
    "tysons-pimmit": "tysons_pimmit",
    "west springfield": "west_springfield",
}

# Sub-column headers inside a 2020 site block; the block's own daily
# figure is the one labelled "Total".
CD_HEADERS = {"cd 8", "cd 10", "cd 11", "c11", "cd 11- herndon", "herndon", "total"}


def norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def site_key(v):
    s = norm(v).lower().rstrip(":").strip()
    return SITE_KEYS.get(s)


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = re.sub(r"[,\s]", "", str(v))
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


# Sheets that look like the ones we want but are derived views, not the
# daily series: "Early Voting Option" sorts before "Early Voting" and
# silently supplied three years of identical wrong numbers before this
# existed.
DECOYS = ("option", "trad", "compare", "for sor", "potential", "curing", "sheet1")


def find_sheet(wb, *needles):
    """Best sheet for a needle group: exact name first, then substring.

    Never returns a decoy sheet. Matching loosely here is how a derived
    summary tab gets parsed as if it were the daily series.
    """
    names = [n for n in wb.sheetnames if not any(d in n.lower() for d in DECOYS)]
    for needle in needles:
        want = needle.lower()
        for name in names:                       # exact wins
            if name.lower().strip() == want:
                return wb[name]
        for name in names:                       # then all-words-present
            if all(w in name.lower() for w in want.split()):
                return wb[name]
    return None


def rows_of(ws):
    return list(ws.iter_rows(values_only=True))


def header_and_total(rows):
    """Locate the header row and the published total row.

    The header is usually labelled "Date", but 2022's early-voting sheet
    leaves that cell empty and starts straight at "Total | <sites…>", so
    a row carrying three or more recognised site names counts as a header
    too.
    """
    header = total = None
    for i, r in enumerate(rows):
        first = norm(r[0]).lower() if r else ""
        joined = " ".join(norm(c).lower() for c in r[:3])
        looks_like_header = (
            first == "date"
            or "date" in joined.split()
            or sum(1 for c in r if site_key(c)) >= 3
        )
        if header is None and looks_like_header:
            header = i
        elif header is not None and total is None and first == "total":
            total = i
    return header, total


def parse_simple(ws, value_col=1):
    """Layout A single-series sheet: {date: value} plus the printed total."""
    rows = rows_of(ws)
    header, total_row = header_and_total(rows)
    if header is None:
        raise SystemExit(f"[{ws.title}] no header row found")
    published = num(rows[total_row][value_col]) if total_row is not None else None
    out = {}
    for r in rows[header + 1:]:
        d = as_date(r[0])
        if d is None:
            continue
        out[d] = num(r[value_col])
    return out, published


def parse_sites_layout_a(ws):
    """2021-2023: one column per site, sites named in the header row."""
    rows = rows_of(ws)
    header, total_row = header_and_total(rows)
    if header is None:
        raise SystemExit(f"[{ws.title}] no header row found")

    cols = {}
    for c, cell in enumerate(rows[header]):
        k = site_key(cell)
        if k:
            cols[k] = c

    published_total = num(rows[total_row][1]) if total_row is not None else None
    published_sites = {
        k: num(rows[total_row][c]) for k, c in cols.items()
    } if total_row is not None else {}

    days = []
    for r in rows[header + 1:]:
        d = as_date(r[0])
        if d is None:
            continue
        days.append((d, num(r[1]), {k: num(r[c]) for k, c in cols.items()}))
    return days, published_total, published_sites


def parse_sites_layout_b(ws):
    """2020: site blocks of CD sub-columns, each closed by a Total column.

    The site name sits one row above the sub-headers, so a block runs from
    one named cell to the next. Within a block we take the column whose
    sub-header is "Total" — the CD splits are not something this site
    models, and summing them would double-count the block total.
    """
    rows = rows_of(ws)
    name_row = sub_row = None
    for i, r in enumerate(rows):
        if any(site_key(c) for c in r):
            name_row = i
            sub_row = i + 1
            break
    if name_row is None:
        raise SystemExit(f"[{ws.title}] no site names found")

    # Where each site block starts.
    starts = [(c, site_key(cell)) for c, cell in enumerate(rows[name_row]) if site_key(cell)]
    cols = {}
    for idx, (c, key) in enumerate(starts):
        end = starts[idx + 1][0] if idx + 1 < len(starts) else len(rows[sub_row])
        for cc in range(c, end):
            if norm(rows[sub_row][cc]).lower() == "total":
                cols[key] = cc
                break

    # County-wide daily total: the column under "Total In Person".
    county_col = 1
    published_total = None
    for c, cell in enumerate(rows[name_row]):
        if "total in person" in norm(cell).lower():
            county_col = c
            published_total = num(rows[sub_row][c])
            break

    days = []
    for r in rows[sub_row + 1:]:
        d = as_date(r[0])
        if d is None:
            continue
        days.append((d, num(r[county_col]), {k: num(r[c]) for k, c in cols.items()}))

    # Each block prints its own site total in the name row, sitting in the
    # same column as the block's "Total" sub-header. Checking these
    # individually is what catches a block boundary landing one column off.
    published_sites = {}
    for key, c in cols.items():
        v = num(rows[name_row][c])
        if v is not None:
            published_sites[key] = v
    return days, published_total, published_sites


def county_column(rows, hdr):
    """Column carrying the county-wide daily series, verified by sum.

    Candidate grand totals are the large bare numbers printed above the
    data; the right column is the one whose dates sum to one of them.
    """
    # Everything above the first dated row. Bounding this by the header
    # index misses totals printed just under the header, which is exactly
    # where 2020's drop-box sheet puts its 85,292.
    first_date = next((i for i, r in enumerate(rows) if as_date(r[0])), len(rows))
    top = rows[:first_date]
    candidates = {
        num(c) for r in top for c in r
        if isinstance(c, (int, float)) and c >= 1000
    }
    width = max(len(r) for r in rows)
    for c in range(1, min(width, 14)):
        total = 0
        for r in rows:
            if as_date(r[0]) is None or c >= len(r):
                continue
            v = num(r[c])
            if v:
                total += v
        if total in candidates:
            return c
    return 1


def check(label, got, want):
    if want is None:
        print(f"    {label}: {got:,} (no published total to check)")
        return
    if got != want:
        raise SystemExit(f"[FAIL] {label}: parsed {got:,}, report publishes {want:,}")
    print(f"    {label}: {got:,} == published {want:,}  ok")


def write_csv(path, header, rows):
    OUT.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)
    print(f"    wrote {path}")


def run(xlsx, cycle, layout, expect):
    """`expect` comes from the workbook's Summary sheet — a *different*
    sheet from the ones parsed here. Checking a daily sheet against its
    own total row proves only that the sheet is internally consistent,
    not that we read the right sheet."""
    print(f"\n=== {cycle}  ({xlsx})")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    prefix = OUT / cycle

    # --- early in person, by site -----------------------------------
    ws = find_sheet(wb, "in-person daily", "early voting", "early in person")
    days, pub_total, pub_sites = (
        parse_sites_layout_b(ws) if layout == "b" else parse_sites_layout_a(ws)
    )
    site_keys = sorted({k for _, _, s in days for k in s})
    header = ["date", "total"] + site_keys
    rows = []
    for d, tot, sites in days:
        rows.append([d.isoformat(), "" if tot is None else tot]
                    + ["" if sites.get(k) is None else sites[k] for k in site_keys])
    got_ip = sum(r[1] for r in rows if r[1] != "")
    check("early in person", got_ip, pub_total)
    if expect.get("inPerson") is not None:
        check("early in person vs Summary", got_ip, expect["inPerson"])
    # Each site column is checked individually — a shifted column still
    # sums correctly county-wide, which is how a real bug once hid.
    for k in site_keys:
        i = header.index(k)
        got = sum(r[i] for r in rows if r[i] != "")
        if pub_sites.get(k) is not None:
            check(f"  site {k}", got, pub_sites[k])
    write_csv(f"{prefix}_early_in_person_by_site.csv", header, rows)

    # --- the three single-series sheets -----------------------------
    for sheet_names, out_name, col_name in [
        (("mailed", "mailout"), "mailed_absentee_ballots", "total_mailed"),
        (("returned by mail", "return by mail", "returned mail"),
         "returned_by_mail", "total_returned"),
        (("returned by dropbox", "return by dropbox", "returned drop box"),
         "returned_by_dropbox", "total_returned_dropbox"),
    ]:
        sh = find_sheet(wb, *sheet_names)
        if sh is None:
            print(f"    (no sheet for {out_name})")
            continue
        # 2020's sheets each put the county series in a different column:
        # "Return by Mail" closes a block of sub-columns with Total at
        # index 5, while "Return by Dropbox" carries it at index 1. Rather
        # than guess by position or by a "Total" header — which lands on a
        # congressional-district column — pick the column whose daily sum
        # equals a grand total printed in the sheet's own top rows.
        vcol = 1
        rws = rows_of(sh)
        hdr, _ = header_and_total(rws)
        if layout == "b":
            vcol = county_column(rws, hdr)
        series, pub = parse_simple(sh, vcol)
        rows = [[d.isoformat(), "" if v is None else v] for d, v in sorted(series.items())]
        got = sum(r[1] for r in rows if r[1] != "")
        check(out_name, got, pub)
        if expect.get(out_name) is not None:
            check(f"{out_name} vs Summary", got, expect[out_name])
        write_csv(f"{prefix}_{out_name}.csv", ["date", col_name], rows)

    # --- mail requesters who voted in person instead ----------------
    sh = find_sheet(wb, "ab voting early instead", "in person instead",
                    "g'rods & surrendered", "surrendered")
    if sh is not None:
        rws = rows_of(sh)
        hdr, _ = header_and_total(rws)
        if hdr is not None:
            series, pub = parse_simple(sh, 1)
            # 2020's sheet counts goldenrods and surrendered separately;
            # its "total" column is the third.
            if layout == "b":
                series, pub = parse_simple(sh, 3)
            rows = [[d.isoformat(), "" if v is None else v] for d, v in sorted(series.items())]
            got = sum(r[1] for r in rows if r[1] != "")
            print(f"    ab in person instead: {got:,}"
                  + (f" (published {pub:,})" if pub else ""))
            write_csv(f"{prefix}_ab_applicants_voted_early_in_person.csv",
                      ["date", "total"], rows)


# Every final report, with the grand totals printed on its own Summary
# sheet. Keeping these in code rather than on the command line means the
# reconciliation is versioned with the parser and re-runs identically.
REPORTS = [
    dict(cycle="fairfax-2020-general", layout="b",
         xlsx="data/sources/fairfax-2020-11-final-ab-daily-report.xlsx",
         # 2020's Summary breaks returns down by congressional district and
         # its column sums fall 42 (mail) and 125 (drop box) short of the
         # daily sheets — ballots the county did not attribute to a
         # district. The daily series is what this site charts, so the
         # daily totals are authoritative here and only drop box, which
         # agrees, is cross-checked.
         inPerson=193596, returned_by_dropbox=85292),
    dict(cycle="fairfax-2021-general", layout="a",
         xlsx="data/sources/fairfax-2021-11-final-ab-daily-report.xlsx",
         inPerson=109764, mailed_absentee_ballots=82239,
         returned_by_mail=48058, returned_by_dropbox=19217),
    dict(cycle="fairfax-2022-general", layout="a",
         xlsx="data/sources/fairfax-2022-11-final-ab-daily-report.xlsx",
         inPerson=82168, mailed_absentee_ballots=76338,
         returned_by_mail=45840, returned_by_dropbox=12346),
    dict(cycle="fairfax-2023-general", layout="a",
         xlsx="data/sources/fairfax-2023-11-final-ab-daily-report.xlsx",
         inPerson=64382, mailed_absentee_ballots=70465,
         returned_by_mail=30240, returned_by_dropbox=6533),
]

KEYS = ("inPerson", "mailed_absentee_ballots",
        "returned_by_mail", "returned_by_dropbox")


def run_all():
    for r in REPORTS:
        run(r["xlsx"], r["cycle"], r["layout"],
            {k: r.get(k) for k in KEYS})


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?")
    ap.add_argument("--cycle")
    ap.add_argument("--all", action="store_true",
                    help="parse every report in REPORTS with its known totals")
    ap.add_argument("--layout", choices=["a", "b"], default="a")
    for k in KEYS:
        ap.add_argument(f"--{k}", type=int, default=None)
    a = ap.parse_args()
    if a.all:
        run_all()
        return
    if not a.xlsx or not a.cycle:
        ap.error("give an xlsx and --cycle, or use --all")
    expect = {
        "inPerson": a.inPerson,
        "mailed_absentee_ballots": a.mailed_absentee_ballots,
        "returned_by_mail": a.returned_by_mail,
        "returned_by_dropbox": a.returned_by_dropbox,
    }
    run(a.xlsx, a.cycle, a.layout, expect)


if __name__ == "__main__":
    main()
